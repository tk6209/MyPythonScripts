# import Workbook
from openpyxl import Workbook
# import load_workbook
from openpyxl import load_workbook


# create Workbook object
wb=Workbook()

# set file path
filepath="C:/Users/vteixeira/Desktop/demo.xlsx"

# save workbook
wb.save(filepath)

#Write
filepath="C:/Users/vteixeira/Desktop/demo.xlsx"

#load wb
wb=load_workbook(filepath)

# select demo.xlsx
sheet=wb.active

# set value for cell A1=1
sheet['A1'] = ("Reports")
sheet['B1'] = ("Data")


# set value for cell B2=2
sheet.cell(row=2, column=1).value = "extvisa"
sheet.cell(row=2, column=2).value = "06/11/2019"

# save workbook
wb.save(filepath)



import 